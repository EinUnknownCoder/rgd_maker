import math
from openpyxl import load_workbook
from pytube import YouTube
from pydub import AudioSegment
import random
from datetime import datetime
import os

# INPUT
length = 0 #Length of the sound file
participants = [] #List of Participants
fair_distribution = False # Every member gets the same number of songs

#SCRIPT
print("Welcome to the RDG Generator!")

print("\nLoading Excel...", end=' ')

wb = load_workbook(filename="songlist.xlsx", read_only=True)
sheet = wb["Songlist"]
already_downloaded = os.listdir("raw")

print("Done!")

print("\nNow Loading songs!")

playlist = []

#for x in range(2, 3):
for x in range(2, sheet.max_row + 1):

    url = sheet.cell(x, 1).value
    title = sheet.cell(x, 2).value
    artist = sheet.cell(x, 3).value
    start = sheet.cell(x, 5).value
    end = sheet.cell(x, 6).value

    file_name = artist + " - " + title + ".mp4"

    if file_name not in already_downloaded:

        print(f"\n{file_name} missing! Downloading... ", end=' ')

        yt = YouTube(url)
        audio_stream = yt.streams.get_audio_only()
        audio_stream.download("raw/", file_name)

        print("Done!")

    else:
        print(f"{file_name} is already downloaded! Skipping...")

    playlist.append(["raw/" + file_name, start, end])

wb.close()

random.shuffle(playlist)

ending_song = AudioSegment.from_mp3("HandsUp.mp3")[20 * 1000:40 * 1000].fade_in(2000).fade_out(2000) + 4

starting_song = ending_song

export = starting_song

timestamps = []
timestamps.append(["0:00", "Intro"])

print("\nCreating Playlist with following order:")

for x in playlist:

    mp4_file_name = x[0].split("/")[1]
    print(mp4_file_name)

    timestamp_minute = str(math.floor(int(export.duration_seconds) / 60))
    timestamp_second = int(export.duration_seconds) % 60
    if(timestamp_second < 10):
        timestamp_second = "0" + str(timestamp_second)
    else:
        timestamp_second = str(timestamp_second)

    timestamp = ":".join([timestamp_minute, timestamp_second])

    timestamps.append([ timestamp , mp4_file_name.rsplit(".", 1)[0] ])

    start = x[1] - 10
    end = x[2] + 2
    song = AudioSegment.from_file(x[0])[start * 1000:end * 1000].fade_in(2000).fade_out(2000)

    export = export + song

timestamp_minute = str(math.floor(int(export.duration_seconds) / 60))
timestamp_second = int(export.duration_seconds) % 60
if(timestamp_second < 10):
    timestamp_second = "0" + str(timestamp_second)
else:
    timestamp_second = str(timestamp_second)

timestamp = ":".join([timestamp_minute, timestamp_second])

timestamps.append([timestamp, "Outro"])

export = export + ending_song

print("\nExporting file...", end=' ')

file_name = "".join([str(datetime.now().strftime("%Y-%m-%d_%H_%M_%S"))]) + ".mp3"

#export.export("export/" + file_name , format="mp3")

print("Done!")

print(f"\nFile {file_name} is created!")

print("Creating Textfile...")

text = ""

for x in timestamps:
    text = text + x[0] + " " + x[1] + "\n"

print(text)

print(f"ffmpeg -loop 1 -framerate 1 -i image.jpg -i {file_name} -map 0:v -map 1:a -r 10 -vf \"scale='iw-mod(iw,2)':'ih-mod(ih,2)',format=yuv420p\" -movflags +faststart -shortest -fflags +shortest -max_interleave_delta 100M {file_name}.mp4")